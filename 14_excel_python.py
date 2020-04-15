import openpyxl

# Reads an excel file
workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

# Reads a specific sheet from the excel file
sheet = workbook['Sheet1']
print(sheet)

# Reads all the sheets from the excel file
sheets = workbook.sheetnames
print(sheets)

# Gets a cell from a sheet
cell = sheet['A1']
print(str(cell.value))
print(str(sheet['B1'].value))
print(str(sheet['C1'].value))

# Other way to get a cell
cell = sheet.cell(row=1, column=2)
print(str(cell.value))

# Iterates a column in excel
for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

# Creates an excel file
workbook = openpyxl.Workbook()

# Creates a sheet
sheet = workbook['Sheet']

# Inserts a value in the sheet
sheet['A1'] = 42
sheet['A2'] = 'Hello'

# Saves an excel file
workbook.save('example2.xlsx')

# Adds new sheet to an excel file
sheet2 = workbook.create_sheet()
sheet2.title = 'My new sheet name'
workbook.save('example3.xlsx')

# Adds a sheet in a specified position
workbook.create_sheet(index=0, title='My other sheet')
workbook.save('example4.xlsx')